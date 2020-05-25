import numpy as np
import pandas as pd
import openpyxl
import csv
import pickle
import time


def timeit(method, full_info=False):
    def timed(*args, **kw):
        t_start = time.time()
        result = method(*args, **kw)
        t_end = time.time()
        if full_info:
            print('{} ({}, {}) {} sec'.format(method, t_end - t_start, args, kw))
            return t_end - t_start
        else:
            return t_end - t_start
        # return

    return timed


def normalization(x, min_value=0, max_value=0):
    if min_value == max_value == 0:
        min_value = np.min(x)
        max_value = np.max(x)
    return (x - min_value) / (max_value - min_value)


def denormalization(x, min_value, max_value):
    return x * (max_value - min_value) + min_value


def train_test_separate(data_set, partition=0.2, shuffle=True):
    if shuffle:
        np.random.shuffle(data_set)
    train_nr = int((1 - partition) * data_set.shape[0])
    return data_set[:train_nr], data_set[train_nr:]


def read_xl(file_name, sheets_number: int = 1, by_cell=False, delimiters: str = ','):
    wb = openpyxl.load_workbook(file_name)
    result_workbook = []
    for sheet_name in wb.sheetnames[:min(sheets_number, len(wb.sheetnames))]:
        ws = wb[sheet_name]
        result_sheet = []
        for row in ws.rows:
            result_row = []
            if by_cell:
                for cell in row:
                    result_row.append(cell.value)
            else:
                result_row = row[0].value.split(delimiters)
            result_sheet.append(result_row)
        result_workbook.append(result_sheet)
    return result_workbook


def read_csv(file_name, delimiter=','):
    data_set = []
    with open(file_name, newline='') as csv_file:
        reader = csv.reader(csv_file, delimiter=delimiter)
        for row in reader:
            data_set.append(row)
    return data_set


def unpickle(file):
    with open(file, 'rb') as fo:
        dict = pickle.load(fo, encoding='bytes')
    return dict


def to_categorial(collumn, unique_values):
    c_dict = {}
    for i, v in enumerate(unique_values):
        c_dict[v] = i
    col = [c_dict[r] for r in collumn]
    return col
